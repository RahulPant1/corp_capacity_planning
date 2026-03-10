"""Demand Analytics Excel Report Generator.

Produces a comprehensive, actionable Excel workbook from the Demand Analytics tab data.
Returns bytes for st.download_button. Follows the same pattern as report_generator.py.
"""

import io
import math
from datetime import datetime
from typing import Optional, List

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from config.defaults import FORECAST_CAPACITY_ALERT_THRESHOLD


# ── Shared openpyxl helpers ───────────────────────────────────────────────────

NAVY_HEX = "1E3A5F"
BLUE_LIGHT_HEX = "EBF5FF"


def _header_style(ws):
    """Navy header row with white bold text. Freeze pane at A2."""
    fill = PatternFill("solid", fgColor=NAVY_HEX)
    font = Font(bold=True, color="FFFFFF")
    align = Alignment(wrap_text=True, vertical="center")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align
    ws.freeze_panes = "A2"


def _autofit(ws, min_w=10, max_w=42):
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value or "")) for cell in col if cell.row <= 200),
            default=min_w,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max(max_len + 2, min_w), max_w
        )


def _risk_fill(tier: str) -> PatternFill:
    colors_map = {"HIGH": "FFCCCC", "MEDIUM": "FFF3CC", "LOW": "CCFFCC"}
    return PatternFill("solid", fgColor=colors_map.get(tier, "FFFFFF"))


def _write_info_row(writer, sheet_name: str, msg: str):
    pd.DataFrame([{"Note": msg}]).to_excel(
        writer, sheet_name=sheet_name, index=False
    )
    ws = writer.sheets[sheet_name]
    _header_style(ws)
    _autofit(ws)


# ── Section writers ───────────────────────────────────────────────────────────

def _write_da_executive_summary(
    writer, daily_df, unit_names, summaries, stf_results, alert_days,
    conflict, breach_data, clusters, rule_config, forecast_months,
    scenario=None,
):
    sheet = "Executive Summary"

    stf_horizon = len(stf_results) if stf_results else 0
    risk_days = len(alert_days) if alert_days else 0
    peak_day_row = max(stf_results, key=lambda r: r.get("expected_seats", 0)) if stf_results else None

    overloaded_days = conflict.get("overloaded_days", []) if conflict else []
    high_risk_units = [
        d["unit_name"] for d in (breach_data or [])
        if d.get("breach_probability", 0) >= 0.20
    ]
    stagger_units = [
        s["unit_name"] for s in (conflict or {}).get("suggestions", [])
    ]

    # Scenario context
    scenario_name = scenario.name if scenario else "Baseline"
    rto_mandate = (
        f"{scenario.params.global_rto_mandate_days}d RTO mandate"
        if scenario and scenario.params and scenario.params.global_rto_mandate_days
        else "Default (attendance-based)"
    )

    # Data period from daily_df
    if daily_df is not None and not daily_df.empty and "date" in daily_df.columns:
        dates = pd.to_datetime(daily_df["date"])
        data_period = f"{dates.min().strftime('%Y-%m-%d')} to {dates.max().strftime('%Y-%m-%d')}"
    else:
        data_period = "N/A"

    rows = [
        ("Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Units Analysed", len(unit_names) if unit_names else 0),
        ("Data Period", data_period),
        ("Scenario (Allocation Rules)", scenario_name),
        ("Allocation Rule", rto_mandate),
        ("", ""),
        ("Short-Term Forecast Days", stf_horizon),
        ("Capacity Risk Days (>90%)", risk_days),
        (
            "Peak Forecast Day",
            f"{peak_day_row['weekday_name']} ({peak_day_row['expected_seats']:,} seats)"
            if peak_day_row else "N/A",
        ),
        ("Overloaded DOW Days", ", ".join(overloaded_days) if overloaded_days else "None"),
        ("Attendance Groups (Clusters)", len(set(c.get("cluster_id") for c in (clusters or [])))),
        ("", ""),
        ("— Action Required —", ""),
        (
            "Units Needing Seat Additions (High Breach Risk ≥20%)",
            ", ".join(high_risk_units) if high_risk_units else "None identified",
        ),
        (
            "Units to Stagger Off Peak Day",
            ", ".join(stagger_units) if stagger_units else "None — load is balanced",
        ),
        ("", ""),
        ("Next Steps", (
            "1. Check Capacity Breach Risk sheet for seat addition priorities. "
            "2. Share Load Balancing sheet with unit managers for stagger coordination. "
            "3. Review Overflow Planning sheet for peak-day flex floor options."
        )),
    ]

    df = pd.DataFrame(rows, columns=["Parameter", "Value"])
    df.to_excel(writer, sheet_name=sheet, index=False)
    ws = writer.sheets[sheet]
    _header_style(ws)

    # Light-blue fill for Action Required block
    action_start = None
    for i, (param, _) in enumerate(rows, start=2):
        if "Action Required" in str(param):
            action_start = i
        if action_start and i >= action_start:
            for col in range(1, 3):
                ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor=BLUE_LIGHT_HEX)

    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 55


def _forecast_action(s) -> str:
    pct = s.get("six_month_change_pct", 0)
    direction = s.get("trend_direction", "")
    change = s.get("six_month_change", 0)
    growth = s.get("suggested_growth_pct", 0)
    if "Growing" in direction and pct > 10:
        return (
            f"Increase allocation by ~{change} seats. "
            f"Apply {growth:.0%} growth in What-If Analysis."
        )
    elif "Declining" in direction and pct < -5:
        return (
            f"Consider releasing {abs(change)} seats. Monitor trends before acting."
        )
    return "No immediate action required. Re-forecast in 3 months."


def _write_da_forecast_summary(writer, summaries, forecast_months):
    sheet = "Forecast Summary"
    if not summaries:
        _write_info_row(writer, sheet, "Insufficient data — need ≥7 days per unit.")
        return

    rows = []
    for s in summaries:
        rows.append({
            "Unit": s.get("unit_name", ""),
            "Current Median": s.get("current_median", 0),
            "Current Peak": s.get("current_peak", 0),
            f"Forecast Median ({forecast_months}m)": s.get("forecasted_median", 0),
            "6M Change (seats)": s.get("six_month_change", 0),
            "6M Change %": f"{s.get('six_month_change_pct', 0):+.1f}%",
            "Trend": s.get("trend_direction", ""),
            "Suggested Growth %": f"{s.get('suggested_growth_pct', 0):.1%}",
            "Recommended Action": _forecast_action(s),
        })

    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name=sheet, index=False)
    ws = writer.sheets[sheet]
    _header_style(ws)

    grey = PatternFill("solid", fgColor="F5F5F5")
    green_light = PatternFill("solid", fgColor="E8F5E9")
    amber_light = PatternFill("solid", fgColor="FFF3CC")

    for i, s in enumerate(summaries, start=2):
        pct = s.get("six_month_change_pct", 0)
        direction = s.get("trend_direction", "")
        if "Growing" in direction and pct > 10:
            fill = green_light
        elif "Declining" in direction and pct < -5:
            fill = amber_light
        elif i % 2 == 0:
            fill = grey
        else:
            fill = None
        if fill:
            for col in range(1, len(rows[0]) + 1):
                ws.cell(row=i, column=col).fill = fill

    _autofit(ws)


def _stf_action(r) -> str:
    pct = r.get("capacity_pct", 0)
    day = r.get("weekday_name", "this day")
    if pct > FORECAST_CAPACITY_ALERT_THRESHOLD:
        return f"Activate overflow floors on {day}. See Overflow Planning sheet."
    elif pct > 0.65:
        return "Monitor — coordinate with Facilities if demand exceeds 80%."
    return "No action required."


def _write_da_short_term_forecast(writer, stf_results, alert_days):
    sheet = "Short-Term Forecast"
    if not stf_results:
        _write_info_row(writer, sheet, "Need ≥7 days of data to generate short-term forecast.")
        return

    rows = []
    for r in stf_results:
        pct = r.get("capacity_pct", 0)
        rows.append({
            "Period": r.get("short_label", ""),
            "Day": r.get("weekday_name", ""),
            "Expected Seats": r.get("expected_seats", 0),
            "Capacity %": f"{pct:.0%}",
            "Risk Flag": "ALERT" if pct > FORECAST_CAPACITY_ALERT_THRESHOLD else ("WATCH" if pct > 0.65 else ""),
            "Recommended Action": _stf_action(r),
        })

    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name=sheet, index=False)
    ws = writer.sheets[sheet]
    _header_style(ws)

    red = PatternFill("solid", fgColor="FFCCCC")
    amber = PatternFill("solid", fgColor="FFF3CC")
    for i, r in enumerate(stf_results, start=2):
        pct = r.get("capacity_pct", 0)
        fill = red if pct > FORECAST_CAPACITY_ALERT_THRESHOLD else (amber if pct > 0.65 else None)
        if fill:
            for col in range(1, len(rows[0]) + 1):
                ws.cell(row=i, column=col).fill = fill

    # Alert Day Summary mini-table
    start_row = len(stf_results) + 4
    ws.cell(row=start_row, column=1).value = "Alert Day Summary"
    ws.cell(row=start_row, column=1).font = Font(bold=True)

    if alert_days:
        hdrs = ["Alert Day", "Expected Seats", "Capacity %"]
        for ci, h in enumerate(hdrs, start=1):
            c = ws.cell(row=start_row + 1, column=ci)
            c.value = h
            c.fill = PatternFill("solid", fgColor=NAVY_HEX)
            c.font = Font(bold=True, color="FFFFFF")
        for ri, r in enumerate(alert_days, start=start_row + 2):
            ws.cell(row=ri, column=1).value = r.get("weekday_name", "")
            ws.cell(row=ri, column=2).value = r.get("expected_seats", 0)
            ws.cell(row=ri, column=3).value = f"{r.get('capacity_pct', 0):.0%}"
            for col in range(1, 4):
                ws.cell(row=ri, column=col).fill = red
    else:
        ws.cell(row=start_row + 1, column=1).value = "No capacity risk days in forecast window."

    _autofit(ws)


def _write_da_dow_patterns(writer, dow_df):
    sheet = "DOW Patterns"
    if dow_df is None or dow_df.empty:
        _write_info_row(writer, sheet, "No DOW data available.")
        return

    # Table A: raw stats
    cols_a = {
        "unit_name": "Unit",
        "day_name": "Day",
        "mean_count": "Mean",
        "median_count": "Median",
        "std_count": "Std Dev",
        "min_count": "Min",
        "max_count": "Max",
    }
    display_a = dow_df[[c for c in cols_a if c in dow_df.columns]].rename(columns=cols_a)
    display_a.to_excel(writer, sheet_name=sheet, index=False)
    ws = writer.sheets[sheet]
    _header_style(ws)

    # Table B: pivot — below Table A
    try:
        day_order = ["Mon", "Tue", "Wed", "Thu", "Fri"]
        pivot = dow_df.pivot_table(
            index="unit_name", columns="day_name",
            values="median_count", aggfunc="sum",
        )
        # Only keep days that exist
        pivot_cols = [d for d in day_order if d in pivot.columns]
        pivot = pivot[pivot_cols].reset_index()
        pivot.rename(columns={"unit_name": "Unit"}, inplace=True)

        start_row = len(display_a) + 4
        # Header for Table B
        ws.cell(row=start_row, column=1).value = "Day-of-Week Median Pivot (seats per day)"
        ws.cell(row=start_row, column=1).font = Font(bold=True)

        # Write pivot headers
        for ci, col_name in enumerate(pivot.columns, start=1):
            c = ws.cell(row=start_row + 1, column=ci)
            c.value = col_name
            c.fill = PatternFill("solid", fgColor=NAVY_HEX)
            c.font = Font(bold=True, color="FFFFFF")

        # Write pivot data and highlight peak day per row
        peak_fill = PatternFill("solid", fgColor="FFCCCC")
        grey = PatternFill("solid", fgColor="F5F5F5")
        day_cols = [c for c in pivot.columns if c in day_order]

        for ri, (_, row_data) in enumerate(pivot.iterrows(), start=start_row + 2):
            for ci, col_name in enumerate(pivot.columns, start=1):
                ws.cell(row=ri, column=ci).value = row_data[col_name]
            if ri % 2 == 0:
                for ci in range(1, len(pivot.columns) + 1):
                    ws.cell(row=ri, column=ci).fill = grey
            # Highlight max day column
            day_values = {col: row_data.get(col, 0) for col in day_cols if col in row_data}
            if day_values:
                peak_col_name = max(day_values, key=day_values.get)
                peak_ci = list(pivot.columns).index(peak_col_name) + 1
                ws.cell(row=ri, column=peak_ci).fill = peak_fill
    except Exception:
        pass  # Pivot is best-effort; Table A always written

    _autofit(ws)


def _breach_action(d) -> str:
    prob = d.get("breach_probability", 0)
    magnitude = d.get("avg_breach_magnitude", 0)
    seats = math.ceil(magnitude / 5) * 5 if magnitude > 0 else 0
    name = d.get("unit_name", "unit")
    if prob >= 0.20:
        return f"HIGH PRIORITY: Add {seats} seats or increase {name} allocation in What-If Analysis."
    elif prob >= 0.10:
        return f"Monitor monthly. Consider adding {seats} seats if trend continues."
    return "Low risk — no immediate action required."


def _write_da_capacity_breach(writer, breach_data):
    sheet = "Capacity Breach Risk"
    if not breach_data:
        _write_info_row(writer, sheet, "No breach data. Run a Policy Simulation in What-If Analysis first.")
        return

    def _tier(prob):
        if prob >= 0.20:
            return "High"
        if prob >= 0.10:
            return "Medium"
        return "Low"

    rows = []
    for d in breach_data:
        prob = d.get("breach_probability", 0)
        magnitude = d.get("avg_breach_magnitude", 0)
        rows.append({
            "Risk Tier": _tier(prob),
            "Unit": d.get("unit_name", ""),
            "Allocated Seats": d.get("allocated_seats", 0),
            "Breach Probability": f"{prob:.0%}",
            "Overflow Days/Month": round(d.get("expected_breach_days_per_month", 0), 1),
            "Avg Overflow (people)": round(magnitude, 1),
            "Seats to Add": math.ceil(magnitude / 5) * 5 if magnitude > 0 else 0,
            "Recommended Action": _breach_action(d),
        })

    # Sort High → Medium → Low
    tier_order = {"High": 0, "Medium": 1, "Low": 2}
    rows.sort(key=lambda r: tier_order.get(r["Risk Tier"], 3))

    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name=sheet, index=False)
    ws = writer.sheets[sheet]
    _header_style(ws)

    for i, r in enumerate(rows, start=2):
        tier = r["Risk Tier"]
        fill_map = {"High": "FFCCCC", "Medium": "FFF3CC", "Low": "CCFFCC"}
        fill = PatternFill("solid", fgColor=fill_map.get(tier, "FFFFFF"))
        for col in range(1, len(rows[0]) + 1):
            ws.cell(row=i, column=col).fill = fill

    _autofit(ws)


def _write_da_load_balancing(writer, conflict, peak_data):
    sheet = "Load Balancing"
    if not conflict and not peak_data:
        _write_info_row(writer, sheet, "No DOW conflict data available.")
        return

    ws_rows_written = 0
    day_loads = (conflict or {}).get("day_loads", {})
    overloaded = set((conflict or {}).get("overloaded_days", []))
    suggestions = (conflict or {}).get("suggestions", [])

    # Sub-table 1: Daily Load Overview
    load_rows = [
        {
            "Day": day,
            "Total Expected Seats": round(seats),
            "Status": "OVERLOADED" if day in overloaded else "Normal",
        }
        for day, seats in day_loads.items()
    ]

    amber = PatternFill("solid", fgColor="FFF3CC")
    grey = PatternFill("solid", fgColor="F5F5F5")

    current_row = 1
    if load_rows:
        load_df = pd.DataFrame(load_rows)
        load_df.to_excel(writer, sheet_name=sheet, index=False, startrow=0)
        ws = writer.sheets[sheet]
        _header_style(ws)
        for i, r in enumerate(load_rows, start=2):
            fill = amber if r["Status"] == "OVERLOADED" else (grey if i % 2 == 0 else None)
            if fill:
                for col in range(1, 4):
                    ws.cell(row=i, column=col).fill = fill
        ws_rows_written = len(load_rows) + 1
    else:
        # Need to create the sheet first
        pd.DataFrame([{"Note": "No daily load data"}]).to_excel(
            writer, sheet_name=sheet, index=False
        )
        ws = writer.sheets[sheet]
        ws_rows_written = 2

    # Sub-table 2: Peak Day per Unit
    start2 = ws_rows_written + 3
    ws.cell(row=start2, column=1).value = "Peak Day per Unit"
    ws.cell(row=start2, column=1).font = Font(bold=True)

    peak_hdrs = ["Unit", "Peak Day", "Peak Median Seats", "Avg Daily Median", "Peak Ratio", "On Overloaded Day?"]
    for ci, h in enumerate(peak_hdrs, start=1):
        c = ws.cell(row=start2 + 1, column=ci)
        c.value = h
        c.fill = PatternFill("solid", fgColor=NAVY_HEX)
        c.font = Font(bold=True, color="FFFFFF")

    peak_data = peak_data or []
    for ri, p in enumerate(peak_data, start=start2 + 2):
        peak_day = p.get("peak_day_name", "")
        on_overloaded = "Yes" if peak_day in overloaded else "No"
        row_vals = [
            p.get("unit_name", ""),
            peak_day,
            round(p.get("peak_day_median", 0)),
            round(p.get("overall_median", 0)),
            f"{p.get('peak_ratio', 1):.2f}x",
            on_overloaded,
        ]
        for ci, v in enumerate(row_vals, start=1):
            cell = ws.cell(row=ri, column=ci)
            cell.value = v
            if on_overloaded == "Yes":
                cell.fill = amber
            elif ri % 2 == 0:
                cell.fill = grey
    ws_rows_written_2 = start2 + 1 + len(peak_data)

    # Sub-table 3: Stagger Suggestions
    start3 = ws_rows_written_2 + 3
    ws.cell(row=start3, column=1).value = "Stagger Suggestions"
    ws.cell(row=start3, column=1).font = Font(bold=True)

    if suggestions:
        sug_hdrs = ["Unit", "Current Peak Day", "Suggested Day", "Est. Load Reduction", "Recommended Action"]
        for ci, h in enumerate(sug_hdrs, start=1):
            c = ws.cell(row=start3 + 1, column=ci)
            c.value = h
            c.fill = PatternFill("solid", fgColor=NAVY_HEX)
            c.font = Font(bold=True, color="FFFFFF")
        for ri, s in enumerate(suggestions, start=start3 + 2):
            reduction = s.get("load_reduction", s.get("est_load_moved", 0))
            action = (
                f"Discuss with {s.get('unit_name', '')} manager: shift RTO anchor from "
                f"{s.get('current_peak_day', '')} to {s.get('suggested_day', '')}. "
                f"Estimated relief: {round(reduction)} fewer seats on peak day."
            )
            row_vals = [
                s.get("unit_name", ""),
                s.get("current_peak_day", ""),
                s.get("suggested_day", ""),
                round(reduction),
                action,
            ]
            for ci, v in enumerate(row_vals, start=1):
                ws.cell(row=ri, column=ci).value = v
    elif overloaded:
        ws.cell(row=start3 + 1, column=1).value = (
            "Overloaded days detected but no specific stagger suggestions generated. "
            "Review unit schedules manually."
        )
    else:
        ws.cell(row=start3 + 1, column=1).value = (
            "Load is balanced across the week — no stagger action required."
        )

    _autofit(ws)


def _write_da_overflow_planning(writer, alert_days, scenario, floors):
    from engine.spatial import get_floor_utilization

    sheet = "Overflow Planning"
    if not (alert_days and getattr(scenario, "floor_assignments", None)
            and getattr(scenario, "allocation_results", None)):
        _write_info_row(
            writer, sheet,
            "Run a Policy Simulation in What-If Analysis first to see floor-level overflow options.",
        )
        return

    floor_util = get_floor_utilization(floors or [], scenario.floor_assignments)
    flex_floors = sorted(
        [f for f in floor_util if f.get("available_seats", 0) > 0],
        key=lambda f: f["available_seats"], reverse=True,
    )
    at_risk = sorted(
        [a for a in scenario.allocation_results if getattr(a, "seat_gap", 0) < 0],
        key=lambda a: getattr(a, "seat_gap", 0),
    )
    overflow_cap = sum(f["available_seats"] for f in flex_floors)

    # Sub-table 1: Capacity Risk Days
    risk_rows = [
        {
            "Alert Day": r.get("weekday_name", ""),
            "Expected Seats": r.get("expected_seats", 0),
            "Capacity %": f"{r.get('capacity_pct', 0):.0%}",
        }
        for r in alert_days
    ]
    risk_df = pd.DataFrame(risk_rows)
    risk_df.to_excel(writer, sheet_name=sheet, index=False, startrow=0)
    ws = writer.sheets[sheet]
    _header_style(ws)
    red = PatternFill("solid", fgColor="FFCCCC")
    for i in range(2, len(risk_rows) + 2):
        for col in range(1, 4):
            ws.cell(row=i, column=col).fill = red
    base_row = len(risk_rows) + 1

    # Sub-table 2: Available Overflow Floors
    start2 = base_row + 3
    ws.cell(row=start2, column=1).value = "Available Overflow Floors (spare seats — sorted by capacity)"
    ws.cell(row=start2, column=1).font = Font(bold=True)

    floor_hdrs = ["Floor", "Tower", "Building", "Spare Seats", "Current Utilization %"]
    for ci, h in enumerate(floor_hdrs, start=1):
        c = ws.cell(row=start2 + 1, column=ci)
        c.value = h
        c.fill = PatternFill("solid", fgColor=NAVY_HEX)
        c.font = Font(bold=True, color="FFFFFF")

    green_light = PatternFill("solid", fgColor="CCFFCC")
    grey = PatternFill("solid", fgColor="F5F5F5")
    for ri, f in enumerate(flex_floors, start=start2 + 2):
        row_vals = [
            f.get("floor_id", ""),
            f.get("tower_id", ""),
            f.get("building_name", ""),
            f.get("available_seats", 0),
            f"{f.get('utilization_pct', 0):.0%}",
        ]
        for ci, v in enumerate(row_vals, start=1):
            cell = ws.cell(row=ri, column=ci)
            cell.value = v
            cell.fill = green_light if ri % 2 == 0 else grey
    base_row2 = start2 + 1 + len(flex_floors)

    # Sub-table 3: Units with Seat Shortfall
    start3 = base_row2 + 3
    ws.cell(row=start3, column=1).value = "Units with Seat Shortfall (demand > allocation — highest priority for overflow)"
    ws.cell(row=start3, column=1).font = Font(bold=True)

    if at_risk:
        short_hdrs = ["Unit", "Allocated Seats", "Demand", "Gap", "Priority Action"]
        for ci, h in enumerate(short_hdrs, start=1):
            c = ws.cell(row=start3 + 1, column=ci)
            c.value = h
            c.fill = PatternFill("solid", fgColor=NAVY_HEX)
            c.font = Font(bold=True, color="FFFFFF")
        for ri, a in enumerate(at_risk, start=start3 + 2):
            gap = getattr(a, "seat_gap", 0)
            needed = abs(gap)
            if overflow_cap >= needed and flex_floors:
                action = f"Direct {needed} staff to {flex_floors[0]['floor_id']} on peak days — sufficient overflow capacity exists."
            else:
                action = f"Insufficient overflow ({overflow_cap} seats available, {needed} needed). Escalate to Facilities for additional flex space."
            row_vals = [
                getattr(a, "unit_name", ""),
                getattr(a, "allocated_seats", 0),
                getattr(a, "effective_demand_seats", 0),
                gap,
                action,
            ]
            for ci, v in enumerate(row_vals, start=1):
                ws.cell(row=ri, column=ci).value = v
                ws.cell(row=ri, column=ci).fill = red
    else:
        ws.cell(row=start3 + 1, column=1).value = (
            "No units have seat shortfalls — all demand is covered by current allocation."
        )

    _autofit(ws)


def _write_da_temporal_clusters(writer, clusters, dow_df):
    sheet = "Temporal Clusters"
    if not clusters:
        _write_info_row(writer, sheet, "Insufficient data for clustering (need ≥2 units).")
        return

    # Build peak day map
    peak_day_map: dict = {}
    if dow_df is not None and not dow_df.empty and "unit_name" in dow_df.columns:
        for unit in dow_df["unit_name"].unique():
            udf = dow_df[dow_df["unit_name"] == unit]
            if not udf.empty:
                peak_row = udf.loc[udf["median_count"].idxmax()]
                peak_day_map[unit] = peak_row.get("day_name", "N/A")

    # Build cluster groups
    cluster_groups: dict = {}
    for r in clusters:
        label = r.get("cluster_label", "")
        cluster_groups.setdefault(label, []).append(r.get("unit_name", ""))

    # Sub-table 1: Unit Cluster Assignments
    rows = []
    for r in clusters:
        label = r.get("cluster_label", "")
        unit = r.get("unit_name", "")
        peers = [m for m in cluster_groups.get(label, []) if m != unit]
        if not peers:
            note = "Sole member of this group — no same-floor constraint."
        else:
            note = (
                f"Do NOT co-locate on the same floor with: {', '.join(peers)}. "
                "They peak simultaneously. Enable Cluster-Diverse Placement in What-If Analysis."
            )
        rows.append({
            "Unit": unit,
            "Cluster ID": r.get("cluster_id", ""),
            "Group Label": label,
            "Group Size": len(cluster_groups.get(label, [])),
            "Peak Day": peak_day_map.get(unit, "N/A"),
            "Planning Note": note,
        })

    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name=sheet, index=False)
    ws = writer.sheets[sheet]
    _header_style(ws)

    palette = ["BBDEFB", "FFCCBC", "C8E6C9", "E1BEE7", "FFF9C4", "B2EBF2"]
    cluster_ids = sorted(set(r.get("cluster_id", 0) for r in clusters))
    cluster_color = {cid: palette[i % len(palette)] for i, cid in enumerate(cluster_ids)}

    grey = PatternFill("solid", fgColor="F5F5F5")
    for i, r in enumerate(clusters, start=2):
        cid = r.get("cluster_id", 0)
        fill = PatternFill("solid", fgColor=cluster_color.get(cid, "FFFFFF"))
        for col in range(1, len(rows[0]) + 1):
            ws.cell(row=i, column=col).fill = fill

    _autofit(ws)


# ── Public entry point ────────────────────────────────────────────────────────

def generate_demand_report(
    daily_df,
    unit_names: list,
    rule_config: dict,
    forecast_months: int = 6,
    summaries: Optional[list] = None,
    stf_results: Optional[list] = None,
    alert_days: Optional[list] = None,
    dow_df=None,
    conflict: Optional[dict] = None,
    peak_data: Optional[list] = None,
    breach_data: Optional[list] = None,
    clusters: Optional[list] = None,
    scenario=None,
    floors: Optional[list] = None,
) -> bytes:
    """Generate a comprehensive Demand Analytics Excel report.

    Returns bytes suitable for st.download_button.
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        try:
            _write_da_executive_summary(
                writer, daily_df, unit_names, summaries, stf_results, alert_days,
                conflict, breach_data or [], clusters, rule_config, forecast_months,
                scenario=scenario,
            )
        except Exception:
            pass
        try:
            _write_da_short_term_forecast(writer, stf_results, alert_days)
        except Exception:
            pass
        try:
            _write_da_dow_patterns(writer, dow_df)
        except Exception:
            pass
        try:
            _write_da_capacity_breach(writer, breach_data or [])
        except Exception:
            pass
        try:
            _write_da_load_balancing(writer, conflict, peak_data)
        except Exception:
            pass
        try:
            _write_da_overflow_planning(writer, alert_days or [], scenario, floors)
        except Exception:
            pass
        try:
            _write_da_temporal_clusters(writer, clusters, dow_df)
        except Exception:
            pass

    return output.getvalue()
