"""Holistic CPG Executive Report — Excel.

Produces a single-workbook 9-sheet Excel report covering ALL domains:
scenario / allocation / floor intelligence / short-term forecast / breach risk /
peak-day operations / demand patterns / unit risk register / scenario comparison.

Returns bytes for st.download_button.
"""

import io
import math
from datetime import datetime
from typing import Optional, List, Dict

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.defaults import (
    RISK_RED_GAP_PCT, RISK_RED_FRAGMENTATION,
    RISK_AMBER_GAP_PCT, RISK_AMBER_FRAGMENTATION,
    FORECAST_CAPACITY_ALERT_THRESHOLD,
)

# ── Palette ───────────────────────────────────────────────────────────────────
NAVY      = "1E3A5F"
NAVY_LITE = "2E5090"
TEAL      = "17A589"
RED_F     = "FFCCCC"
AMBER_F   = "FFF3CC"
GREEN_F   = "D5F5E3"
BLUE_F    = "EBF5FF"
GREY_F    = "F5F5F5"
WHITE_F   = "FFFFFF"

# ── Shared helpers ────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=10) -> Font:
    return Font(bold=bold, color=color, size=size)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _header_row(ws):
    """Navy header, white bold text, freeze pane."""
    for cell in ws[1]:
        cell.fill = _fill(NAVY)
        cell.font = _font(bold=True, color="FFFFFF", size=10)
        cell.alignment = _center()
    ws.freeze_panes = "A2"


def _autofit(ws, min_w=10, max_w=50):
    for col in ws.columns:
        mx = max(
            (len(str(cell.value or "")) for cell in col if cell.row <= 300),
            default=min_w,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(mx + 2, min_w), max_w)


def _risk_level(gap_pct: float, frag: float) -> str:
    if gap_pct < RISK_RED_GAP_PCT or frag > RISK_RED_FRAGMENTATION:
        return "RED"
    if gap_pct < RISK_AMBER_GAP_PCT or frag > RISK_AMBER_FRAGMENTATION:
        return "AMBER"
    return "GREEN"


def _risk_fill(level: str) -> PatternFill:
    return _fill({"RED": RED_F, "AMBER": AMBER_F, "GREEN": GREEN_F}.get(level, WHITE_F))


def _breach_tier(prob: float) -> str:
    if prob >= 0.20:
        return "HIGH"
    if prob >= 0.10:
        return "MEDIUM"
    return "LOW"


def _breach_fill(tier: str) -> PatternFill:
    return _fill({"HIGH": RED_F, "MEDIUM": AMBER_F, "LOW": GREEN_F}.get(tier, WHITE_F))


def _write_df(writer, sheet_name: str, df: pd.DataFrame, color_col: str = None,
              color_fn=None, note_if_empty: str = None):
    """Write a DataFrame to an Excel sheet with navy header + autofit + optional row coloring."""
    if df is None or df.empty:
        if note_if_empty:
            pd.DataFrame([{"Note": note_if_empty}]).to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            _header_row(ws)
            _autofit(ws)
        return
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]
    _header_row(ws)
    if color_col and color_fn and color_col in df.columns:
        col_idx = list(df.columns).index(color_col) + 1
        for row_idx, val in enumerate(df[color_col], start=2):
            bg = color_fn(str(val))
            for c in ws.iter_cols(min_col=1, max_col=ws.max_column,
                                   min_row=row_idx, max_row=row_idx):
                for cell in c:
                    cell.fill = bg
            # Bold the color column cell
            ws.cell(row=row_idx, column=col_idx).font = _font(bold=True, size=10)
    _autofit(ws)


# ── Sheet writers ─────────────────────────────────────────────────────────────

def _write_cover(writer, scenario, floors, units, rule_config, has_daily):
    ws = writer.book.create_sheet("Cover", 0)

    # Title band
    ws.merge_cells("A1:F3")
    title_cell = ws["A1"]
    title_cell.value = "CPG Workforce Seat Intelligence Report"
    title_cell.fill = _fill(NAVY)
    title_cell.font = Font(bold=True, color="FFFFFF", size=20)
    title_cell.alignment = _center()
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 30

    # Scenario meta band
    ws.merge_cells("A4:F4")
    meta = ws["A4"]
    report_ts = datetime.now().strftime("%B %d, %Y  %H:%M")
    meta.value = f"Scenario: {scenario.name}  ·  Type: {scenario.scenario_type}  ·  Generated: {report_ts}"
    meta.fill = _fill(NAVY_LITE)
    meta.font = Font(bold=True, color="FFFFFF", size=11)
    meta.alignment = _center()
    ws.row_dimensions[4].height = 22

    # About section
    about_rows = [
        (6, "About This Report"),
        (7, "This workbook is produced by the CPG Seat Planning & Scenario Intelligence Platform."),
        (8, "It consolidates scenario simulation, short-term demand forecasting, floor space intelligence,"),
        (9, "and unit-level risk assessment into a single executive-ready document."),
    ]
    for r, text in about_rows:
        ws.merge_cells(f"A{r}:F{r}")
        cell = ws[f"A{r}"]
        cell.value = text
        cell.alignment = _left()
        if r == 6:
            cell.font = Font(bold=True, color=NAVY, size=13)
        else:
            cell.font = Font(size=10, color="333333")
        ws.row_dimensions[r].height = 16

    # CPG Use Cases
    ws.merge_cells("A11:F11")
    ws["A11"].value = "How CPG Can Use This Report"
    ws["A11"].font = Font(bold=True, color=NAVY, size=12)
    ws["A11"].alignment = _left()

    use_cases = [
        ("Annual Seat Planning",      "Validate seat allocation against projected headcount growth under different RTO scenarios"),
        ("Capacity Risk Management",  "Identify units at risk of overflow before they happen using short-term demand forecasts"),
        ("Floor Consolidation",        "Spot under-utilised floors for sublease, renovation, or hot-desk conversion"),
        ("RTO Policy Assessment",     "Compare attendance-based vs mandate-driven seat demand to right-size provisioning"),
        ("Peak Day Operations",       "Pre-position overflow floors and stagger team schedules on forecast-heavy days"),
        ("Scenario Planning",         "Compare multiple allocation + RTO policy combinations to find the optimal configuration"),
    ]
    headers = ["CPG Use Case", "Description"]
    ws.append([])  # row 12 blank
    ws.append(headers)  # row 13
    for cell in ws[13]:
        cell.fill = _fill(NAVY)
        cell.font = _font(bold=True, color="FFFFFF")
        cell.alignment = _center()
    for i, (uc, desc) in enumerate(use_cases, start=14):
        ws.append([uc, desc])
        ws[f"A{i}"].font = _font(bold=True, size=10)
        ws[f"A{i}"].fill = _fill(BLUE_F if i % 2 == 0 else WHITE_F)
        ws[f"B{i}"].fill = _fill(BLUE_F if i % 2 == 0 else WHITE_F)
        ws[f"B{i}"].alignment = _left()

    # Table of contents
    toc_start = 22
    ws.merge_cells(f"A{toc_start}:F{toc_start}")
    ws[f"A{toc_start}"].value = "Workbook Contents"
    ws[f"A{toc_start}"].font = Font(bold=True, color=NAVY, size=12)
    ws[f"A{toc_start}"].alignment = _left()

    toc = [
        ("1", "Cover",                     "Application overview and CPG use cases"),
        ("2", "Executive Summary",          "All key KPIs from every domain on one page"),
        ("3", "Scenario & Allocation",      "Per-unit demand, allocated seats, gap and risk level"),
        ("4", "Floor Intelligence",         "Floor utilisation and consolidation opportunities"),
        ("5", "Short-Term Forecast & Risk", "Daily capacity outlook (next 10 days) + breach probability"),
        ("6", "Peak Day Operations",        "Load balancing, stagger suggestions, overflow floors"),
        ("7", "Demand Patterns",            "Day-of-week attendance patterns and attendance clusters"),
        ("8", "Unit Risk Register",         "Comprehensive per-unit risk combining all dimensions"),
        ("9", "Scenario Comparison",        "Ranked scenario variants (if matrix was run)"),
    ]
    ws.append([])
    ws.append(["#", "Sheet", "Contents"])
    for cell in ws[toc_start + 2]:
        if cell.value:
            cell.fill = _fill(TEAL)
            cell.font = _font(bold=True, color="FFFFFF")
            cell.alignment = _center()
    for num, sheet, desc in toc:
        ws.append([num, sheet, desc])

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 56
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 8
    ws.sheet_view.showGridLines = False


def _write_executive_summary(writer, scenario, floors, units, rule_config,
                              stf_results, alert_days, breach_data, clusters,
                              conflict, matrix_results, has_daily):
    supply = sum(f.total_seats for f in floors) if floors else 0
    allocs = scenario.allocation_results or []
    demand = sum(a.effective_demand_seats for a in allocs)
    allocated = sum(a.allocated_seats for a in allocs)
    at_risk = sum(1 for a in allocs if a.seat_gap < 0)
    n_units = len(allocs)
    gap = supply - demand
    util_pct = demand / supply if supply else 0

    red_n = sum(1 for a in allocs if _risk_level(
        a.seat_gap / a.effective_demand_seats if a.effective_demand_seats else 0,
        a.fragmentation_score) == "RED")
    amber_n = sum(1 for a in allocs if _risk_level(
        a.seat_gap / a.effective_demand_seats if a.effective_demand_seats else 0,
        a.fragmentation_score) == "AMBER")

    stf_horizon = len(set(r.get("date") for r in (stf_results or [])))
    risk_days_count = len(alert_days) if alert_days else 0
    high_breach_units = [d["unit_name"] for d in (breach_data or []) if d.get("breach_probability", 0) >= 0.20]
    cluster_count = len(set(c.get("cluster_id") for c in (clusters or [])))
    overloaded_days = (conflict or {}).get("overloaded_days", [])

    rows = [
        ["─── Scenario Context ───", ""],
        ["Active Scenario", scenario.name],
        ["Scenario Type", scenario.scenario_type],
        ["Planning Horizon", f"{scenario.planning_horizon_months} months"],
        ["RTO Mandate", f"{scenario.params.global_rto_mandate_days}d / week" if scenario.params.global_rto_mandate_days else "None (attendance-based)"],
        ["Last Simulation Run", scenario.last_run_at.strftime("%Y-%m-%d %H:%M") if scenario.last_run_at else "Not run"],
        ["Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["", ""],
        ["─── Capacity & Allocation ───", ""],
        ["Total Seat Supply", f"{supply:,}"],
        ["Total Projected Demand", f"{demand:,}"],
        ["Total Seats Allocated", f"{allocated:,}"],
        ["Supply Headroom (Supply − Demand)", f"{gap:+,} seats"],
        ["Capacity Utilisation", f"{util_pct:.0%}"],
        ["Units at Individual Risk", f"{at_risk} of {n_units}"],
        ["RED Risk Units (Critical)", str(red_n)],
        ["AMBER Risk Units (Monitor)", str(amber_n)],
        ["GREEN Risk Units (Healthy)", str(n_units - red_n - amber_n)],
        ["", ""],
        ["─── Short-Term Demand Outlook ───", ""],
        ["Forecast Horizon", f"{stf_horizon} days" if has_daily else "N/A — load daily attendance data"],
        ["Capacity Breach Days (>90%)", str(risk_days_count) if has_daily else "N/A"],
        ["High Breach-Risk Units (≥20% days)", ", ".join(high_breach_units) if high_breach_units else "None" if has_daily else "N/A"],
        ["Overloaded DOW Days", ", ".join(overloaded_days) if overloaded_days else "None" if has_daily else "N/A"],
        ["Attendance Clusters", str(cluster_count) if has_daily else "N/A"],
        ["", ""],
        ["─── Recommended Actions ───", ""],
    ]

    # Auto-generate top 3 actions
    actions = []
    if red_n > 0:
        worst = sorted(allocs, key=lambda a: a.seat_gap)[:2]
        names = ", ".join(a.unit_name for a in worst)
        actions.append(f"URGENT — Add seats for {names} (RED risk, seat shortfall).")
    if risk_days_count > 0 and has_daily:
        actions.append(f"Coordinate overflow floors for {risk_days_count} high-demand day(s) in the next {stf_horizon} days.")
    if high_breach_units and has_daily:
        actions.append(f"Review allocation for {', '.join(high_breach_units[:2])} — historically overflow ≥20% of days.")
    if overloaded_days and has_daily:
        actions.append(f"Apply stagger scheduling on {', '.join(overloaded_days)} — these DOWs consistently exceed capacity.")
    under_util = [f for f in (floors or []) if hasattr(f, "total_seats") and f.total_seats > 0]
    if amber_n > 3:
        actions.append("Run Scenario Comparison Matrix in What-If Analysis to find a more efficient policy.")
    if not actions:
        actions.append("Capacity and demand are well-balanced. Continue monitoring short-term forecasts weekly.")

    for i, act in enumerate(actions[:5], start=1):
        rows.append([f"Action {i}", act])

    if matrix_results:
        best = next((r for r in matrix_results if r.get("rank") == 1), None)
        if best:
            rows.append(["", ""])
            rows.append(["─── Scenario Comparison ───", ""])
            rows.append(["Best Ranked Scenario", f"Rank #1 — RTO {best.get('rto_mandate')}d | {best.get('objective','')} | Headroom: {best.get('headroom','?')}"])

    df = pd.DataFrame(rows, columns=["Metric", "Value"])
    df.to_excel(writer, sheet_name="Executive Summary", index=False)
    ws = writer.sheets["Executive Summary"]
    _header_row(ws)

    # Style section header rows and risk rows
    for row_idx in range(2, ws.max_row + 1):
        label = ws.cell(row=row_idx, column=1).value or ""
        val = ws.cell(row=row_idx, column=2).value or ""
        if str(label).startswith("───"):
            ws.cell(row=row_idx, column=1).fill = _fill(NAVY_LITE)
            ws.cell(row=row_idx, column=1).font = _font(bold=True, color="FFFFFF")
            ws.cell(row=row_idx, column=2).fill = _fill(NAVY_LITE)
        elif "RED" in str(label):
            ws.cell(row=row_idx, column=2).fill = _fill(RED_F)
        elif "AMBER" in str(label):
            ws.cell(row=row_idx, column=2).fill = _fill(AMBER_F)
        elif "GREEN" in str(label) and "Healthy" in str(label):
            ws.cell(row=row_idx, column=2).fill = _fill(GREEN_F)
        elif str(label).startswith("Action"):
            ws.cell(row=row_idx, column=1).font = _font(bold=True, size=10)
            ws.cell(row=row_idx, column=2).fill = _fill(BLUE_F)

    _autofit(ws)


def _write_scenario_allocation(writer, scenario, units, attendance_map, rule_config):
    from engine.allocation_engine import compute_rto_alerts
    allocs = scenario.allocation_results or []
    if not allocs:
        _write_df(writer, "Scenario & Allocation", None,
                  note_if_empty="No simulation results. Run Policy Simulation in What-If Analysis.")
        return

    rto_alerts = compute_rto_alerts(allocs, units, attendance_map, rule_config)
    rto_map = {r["unit_name"]: r for r in rto_alerts}
    unit_map = {u.unit_name: u for u in units}

    rows = []
    for a in sorted(allocs, key=lambda x: x.seat_gap):
        u = unit_map.get(a.unit_name)
        ra = rto_map.get(a.unit_name)
        gap_pct = a.seat_gap / a.effective_demand_seats if a.effective_demand_seats > 0 else 0
        risk = _risk_level(gap_pct, a.fragmentation_score)
        rows.append({
            "Risk": risk,
            "Unit": a.unit_name,
            "Priority": (u.business_priority or "—") if u else "—",
            "Current HC": u.current_total_hc if u else "—",
            "Alloc %": f"{a.recommended_alloc_pct:.0%}",
            "Demand Seats": a.effective_demand_seats,
            "Allocated Seats": a.allocated_seats,
            "Gap (Alloc − Demand)": f"{a.seat_gap:+,}",
            "RTO Need (seats)": ra["expected_seats"] if ra else "N/A",
            "RTO Status": ra["status"] if ra else "N/A",
            "Fragmentation": f"{a.fragmentation_score:.2f}",
            "Overridden": "Yes" if a.is_overridden else "No",
            "Recommended Action": (
                "Add seats immediately" if risk == "RED" and a.seat_gap < 0 else
                "Monitor — approaching limit" if risk == "AMBER" else
                "Right-size if surplus persists" if a.seat_gap > 20 else
                "No action needed"
            ),
        })

    df = pd.DataFrame(rows)
    _write_df(writer, "Scenario & Allocation", df, color_col="Risk",
              color_fn=lambda v: _risk_fill(v))


def _write_floor_intelligence(writer, scenario, floors):
    from engine.spatial import get_floor_utilization
    floor_assignments = scenario.floor_assignments or []
    if not floors or not floor_assignments:
        _write_df(writer, "Floor Intelligence", None,
                  note_if_empty="No floor assignment data. Run Policy Simulation first.")
        return

    util = get_floor_utilization(floors, floor_assignments)
    rows = []
    for f in sorted(util, key=lambda x: (x.get("tower_id", ""), x.get("floor_number", 0))):
        total = f.get("total_seats", 0)
        occupied = f.get("used_seats", 0)
        available = f.get("available_seats", 0)
        util_pct = f.get("utilization_pct", occupied / total if total > 0 else 0)
        rows.append({
            "Tower": f.get("tower_id", "—"),
            "Floor": f.get("floor_number", "—"),
            "Floor ID": f.get("floor_id", "—"),
            "Total Seats": total,
            "Occupied Seats": occupied,
            "Available Seats": available,
            "Utilisation %": f"{util_pct:.0%}",
            "Status": (
                "Over-allocated" if occupied > total else
                "Near capacity (>90%)" if util_pct >= 0.90 else
                "Healthy (60–90%)" if util_pct >= 0.60 else
                "Under-utilised (<60%)" if util_pct > 0 else
                "Empty"
            ),
        })

    df = pd.DataFrame(rows)

    def _status_fill(status: str) -> PatternFill:
        if "Over" in status or "Near" in status:
            return _fill(AMBER_F)
        if "Under" in status or "Empty" in status:
            return _fill(BLUE_F)
        return _fill(GREEN_F)

    _write_df(writer, "Floor Intelligence", df, color_col="Status",
              color_fn=_status_fill)

    # Consolidation opportunities: floors below 40% utilisation
    low_util = [f for f in util if 0 < f.get("utilization_pct", 0) < 0.40]
    if low_util:
        ws = writer.sheets["Floor Intelligence"]
        gap = ws.max_row + 2
        ws.cell(row=gap, column=1).value = "Consolidation Opportunities (< 40% utilised)"
        ws.cell(row=gap, column=1).font = _font(bold=True, size=11, color=NAVY)
        gap += 1
        for cell_col, hdr in enumerate(["Floor ID", "Tower", "Floor #", "Utilisation %", "Available Seats", "Recommendation"], start=1):
            ws.cell(row=gap, column=cell_col).value = hdr
            ws.cell(row=gap, column=cell_col).fill = _fill(TEAL)
            ws.cell(row=gap, column=cell_col).font = _font(bold=True, color="FFFFFF")
        for f in sorted(low_util, key=lambda x: x.get("utilization_pct", 0)):
            gap += 1
            util_pct = f.get("utilization_pct", 0)
            ws.cell(row=gap, column=1).value = f.get("floor_id", "—")
            ws.cell(row=gap, column=2).value = f.get("tower_id", "—")
            ws.cell(row=gap, column=3).value = f.get("floor_number", "—")
            ws.cell(row=gap, column=4).value = f"{util_pct:.0%}"
            ws.cell(row=gap, column=5).value = f.get("available_seats", "—")
            ws.cell(row=gap, column=6).value = (
                "Consider sublease or decommission" if util_pct < 0.20 else
                "Candidate for hot-desk or flex space"
            )


def _write_stf_and_breach(writer, scenario, stf_results, alert_days, breach_data, has_daily):
    if not has_daily:
        _write_df(writer, "Short-Term Forecast & Risk", None,
                  note_if_empty="Load daily attendance data in Admin tab to enable this section.")
        return

    # Part 1: Daily STF forecast — aggregate per-unit rows to daily totals
    stf_rows = []
    supply_total = sum(a.allocated_seats for a in (scenario.allocation_results or []))
    # stf_results is per-unit per-day; aggregate to daily totals first
    daily_agg: dict = {}
    for r in (stf_results or []):
        date_key = str(r.get("date", ""))[:10]
        if date_key not in daily_agg:
            daily_agg[date_key] = {"weekday_name": r.get("weekday_name", ""), "expected_seats": 0}
        daily_agg[date_key]["expected_seats"] += r.get("expected_seats", 0)
    for date_key in sorted(daily_agg):
        expected = daily_agg[date_key]["expected_seats"]
        cap_pct = expected / supply_total if supply_total > 0 else 0
        stf_rows.append({
            "Date": date_key,
            "Day": daily_agg[date_key]["weekday_name"],
            "Expected Seats": expected,
            "Allocated Capacity": supply_total,
            "Capacity %": f"{cap_pct:.0%}",
            "Alert": "⚠ Over 90%" if cap_pct > FORECAST_CAPACITY_ALERT_THRESHOLD else "",
        })

    # Part 2: Breach risk per unit
    breach_rows = []
    for d in (breach_data or []):
        prob = d.get("breach_probability", 0)
        tier = _breach_tier(prob)
        mag = d.get("avg_breach_magnitude", 0)
        seats_to_add = int(math.ceil(mag / 5) * 5) if mag > 0 else 0
        breach_rows.append({
            "Breach Risk": tier,
            "Unit": d["unit_name"],
            "Allocated Seats": d.get("allocated_seats", "—"),
            "% Days Overflow": f"{prob:.0%}",
            "Overflow Days/Month": f"~{d.get('expected_breach_days_per_month', 0):.0f}",
            "Avg Overflow People": f"+{mag:.0f}" if mag > 0 else "0",
            "Seats to Add (Fix)": f"+{seats_to_add}" if seats_to_add > 0 else "—",
            "Recommended Action": (
                "Add seats urgently" if tier == "HIGH" else
                "Monitor and plan buffer" if tier == "MEDIUM" else
                "No action needed"
            ),
        })

    # Write STF sheet
    df_stf = pd.DataFrame(stf_rows) if stf_rows else None

    def _alert_fill(alert_val: str) -> PatternFill:
        return _fill(AMBER_F) if "Over" in str(alert_val) else _fill(WHITE_F)

    _write_df(writer, "Short-Term Forecast & Risk", df_stf,
              color_col="Alert", color_fn=_alert_fill,
              note_if_empty="No short-term forecast data available.")

    # Append breach table below STF
    if breach_rows:
        ws = writer.sheets["Short-Term Forecast & Risk"]
        gap = ws.max_row + 2
        ws.cell(row=gap, column=1).value = "Capacity Breach Risk by Unit (Historical)"
        ws.cell(row=gap, column=1).font = _font(bold=True, size=11, color=NAVY)
        gap += 1
        df_breach = pd.DataFrame(breach_rows).sort_values(
            "Breach Risk", key=lambda s: s.map({"HIGH": 0, "MEDIUM": 1, "LOW": 2})
        )
        for col_i, col_name in enumerate(df_breach.columns, start=1):
            ws.cell(row=gap, column=col_i).value = col_name
            ws.cell(row=gap, column=col_i).fill = _fill(NAVY)
            ws.cell(row=gap, column=col_i).font = _font(bold=True, color="FFFFFF")
        for row_i, row_data in enumerate(df_breach.itertuples(index=False), start=gap + 1):
            tier_val = row_data[0]
            bg = _breach_fill(str(tier_val))
            for col_i, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row_i, column=col_i)
                cell.value = val
                cell.fill = bg
        _autofit(ws)


def _write_peak_day_ops(writer, scenario, peak_data, conflict, floors, has_daily, clusters):
    if not has_daily:
        _write_df(writer, "Peak Day Operations", None,
                  note_if_empty="Load daily attendance data in Admin tab to enable this section.")
        return

    cluster_map = {c["unit_name"]: c.get("cluster_id") for c in (clusters or [])} if clusters else {}

    # Per-unit peak day table
    peak_rows = []
    for p in (peak_data or []):
        name = p.get("unit_name", "")
        cid = cluster_map.get(name)
        peak_rows.append({
            "Unit": name,
            "Cluster Group": f"Group {cid}" if cid is not None else "—",
            "Peak Day": p.get("peak_dow_name", "—"),
            "Peak Expected Seats": p.get("peak_seats", "—"),
            "Secondary Peak Day": p.get("secondary_dow_name", "—"),
        })

    df_peak = pd.DataFrame(peak_rows) if peak_rows else None
    _write_df(writer, "Peak Day Operations", df_peak,
              note_if_empty="No peak day data available.")

    ws = writer.sheets["Peak Day Operations"]

    # Stagger suggestions
    suggestions = (conflict or {}).get("suggestions", [])
    overloaded = (conflict or {}).get("overloaded_days", [])
    if suggestions or overloaded:
        gap = ws.max_row + 2
        ws.cell(row=gap, column=1).value = "Stagger & Load Balancing Suggestions"
        ws.cell(row=gap, column=1).font = _font(bold=True, size=11, color=NAVY)
        if overloaded:
            gap += 1
            ws.cell(row=gap, column=1).value = f"Overloaded DOW days: {', '.join(overloaded)}"
            ws.cell(row=gap, column=1).fill = _fill(AMBER_F)
        gap += 1
        sug_headers = ["Unit", "Current Peak DOW", "Suggested DOW", "Est. Load Reduction", "Cluster Note"]
        for ci, h in enumerate(sug_headers, start=1):
            ws.cell(row=gap, column=ci).value = h
            ws.cell(row=gap, column=ci).fill = _fill(TEAL)
            ws.cell(row=gap, column=ci).font = _font(bold=True, color="FFFFFF")
        for s in suggestions:
            gap += 1
            name = s.get("unit_name", "")
            cid = cluster_map.get(name)
            same_cluster_peers = [
                p.get("unit_name", "") for p in (peak_data or [])
                if cluster_map.get(p.get("unit_name", "")) == cid
                and p.get("unit_name") != name
                and p.get("peak_dow_name") == s.get("current_peak_dow")
            ] if cid is not None else []
            ws.cell(row=gap, column=1).value = name
            ws.cell(row=gap, column=2).value = s.get("current_peak_dow", "—")
            ws.cell(row=gap, column=3).value = s.get("suggested_dow", "—")
            ws.cell(row=gap, column=4).value = s.get("estimated_reduction", "—")
            ws.cell(row=gap, column=5).value = (
                f"Same cluster as: {', '.join(same_cluster_peers[:2])} — prioritise stagger"
                if same_cluster_peers else "Different cluster — lower co-peak risk"
            )

    # Overflow floors
    floor_assignments = scenario.floor_assignments or []
    if floors and floor_assignments:
        from engine.spatial import get_floor_utilization
        util = get_floor_utilization(floors, floor_assignments)
        flex_floors = [f for f in util if f.get("available_seats", 0) > 0]
        if flex_floors:
            gap = ws.max_row + 2
            ws.cell(row=gap, column=1).value = "Available Overflow Floors"
            ws.cell(row=gap, column=1).font = _font(bold=True, size=11, color=NAVY)
            gap += 1
            of_headers = ["Floor ID", "Tower", "Floor #", "Total Seats", "Available Seats"]
            for ci, h in enumerate(of_headers, start=1):
                ws.cell(row=gap, column=ci).value = h
                ws.cell(row=gap, column=ci).fill = _fill(TEAL)
                ws.cell(row=gap, column=ci).font = _font(bold=True, color="FFFFFF")
            for ff in sorted(flex_floors, key=lambda x: -x.get("available_seats", 0))[:10]:
                gap += 1
                ws.cell(row=gap, column=1).value = ff.get("floor_id", "—")
                ws.cell(row=gap, column=2).value = ff.get("tower_id", "—")
                ws.cell(row=gap, column=3).value = ff.get("floor_number", "—")
                ws.cell(row=gap, column=4).value = ff.get("total_seats", "—")
                ws.cell(row=gap, column=5).value = ff.get("available_seats", "—")
                ws.cell(row=gap, column=5).fill = _fill(GREEN_F)

    _autofit(ws)


def _write_demand_patterns(writer, dow_df, clusters, has_daily):
    if not has_daily or dow_df is None or dow_df.empty:
        _write_df(writer, "Demand Patterns", None,
                  note_if_empty="Load daily attendance data in Admin tab to enable this section.")
        return

    # DOW pivot
    DAY_ORDER = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
    try:
        pivot = dow_df.pivot_table(
            index="unit_name", columns="day_name",
            values="median_count", aggfunc="first",
        )[DAY_ORDER]
        pivot.columns.name = None
        pivot = pivot.reset_index().rename(columns={"unit_name": "Unit"})
        pivot.to_excel(writer, sheet_name="Demand Patterns", index=False)
        ws = writer.sheets["Demand Patterns"]
        _header_row(ws)

        # Highlight peak day per unit (row-wise max)
        data_cols = [c for c in pivot.columns if c != "Unit"]
        for row_idx in range(2, ws.max_row + 1):
            vals = {}
            for col_idx in range(2, len(data_cols) + 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    vals[col_idx] = float(cell.value)
                except (TypeError, ValueError):
                    pass
            if vals:
                peak_col = max(vals, key=vals.get)
                ws.cell(row=row_idx, column=peak_col).fill = _fill(AMBER_F)
                ws.cell(row=row_idx, column=peak_col).font = _font(bold=True)

        ws.cell(row=ws.max_row + 1, column=1).value = "◀ Highlighted cell = peak attendance day for that unit"
        ws.cell(row=ws.max_row, column=1).font = Font(italic=True, size=8, color="888888")
    except Exception:
        pass

    # Temporal clusters below pivot
    if clusters:
        ws = writer.sheets["Demand Patterns"]
        gap = ws.max_row + 2
        ws.cell(row=gap, column=1).value = "Attendance Cluster Groups"
        ws.cell(row=gap, column=1).font = _font(bold=True, size=11, color=NAVY)
        gap += 1
        cl_headers = ["Unit", "Cluster Group", "Group Size", "Peak DOW", "Co-Peak Risk Note"]
        for ci, h in enumerate(cl_headers, start=1):
            ws.cell(row=gap, column=ci).value = h
            ws.cell(row=gap, column=ci).fill = _fill(NAVY)
            ws.cell(row=gap, column=ci).font = _font(bold=True, color="FFFFFF")

        # Build cluster peer map
        cluster_peers: Dict[int, List[str]] = {}
        for c in clusters:
            cid = c.get("cluster_id")
            if cid is not None:
                cluster_peers.setdefault(cid, []).append(c.get("unit_name", ""))

        CLUSTER_COLORS = ["D6EAF8", "D5F5E3", "FEF9E7", "FADBD8", "E8DAEF"]
        for c in sorted(clusters, key=lambda x: (x.get("cluster_id", 99), x.get("unit_name", ""))):
            gap += 1
            cid = c.get("cluster_id", 0)
            peers = cluster_peers.get(cid, [])
            bg = CLUSTER_COLORS[cid % len(CLUSTER_COLORS)]
            for ci in range(1, 6):
                ws.cell(row=gap, column=ci).fill = _fill(bg)
            ws.cell(row=gap, column=1).value = c.get("unit_name", "—")
            ws.cell(row=gap, column=2).value = f"Group {cid}"
            ws.cell(row=gap, column=3).value = len(peers)
            ws.cell(row=gap, column=4).value = c.get("peak_dow", "—")
            ws.cell(row=gap, column=5).value = (
                f"Correlated peak with: {', '.join(p for p in peers if p != c.get('unit_name'))}"
                if len(peers) > 1 else "No strong correlations"
            )

        _autofit(ws)


def _write_unit_risk_register(writer, scenario, units, attendance_map, rule_config,
                               breach_data, stf_results, clusters, has_daily):
    from engine.allocation_engine import compute_rto_alerts
    allocs = scenario.allocation_results or []
    if not allocs:
        _write_df(writer, "Unit Risk Register", None,
                  note_if_empty="No simulation results. Run Policy Simulation first.")
        return

    unit_map = {u.unit_name: u for u in units}
    rto_alerts = compute_rto_alerts(allocs, units, attendance_map, rule_config)
    rto_map = {r["unit_name"]: r for r in rto_alerts}
    breach_map = {d["unit_name"]: d for d in (breach_data or [])}
    cluster_map = {c["unit_name"]: c.get("cluster_id") for c in (clusters or [])}

    # STF peak per unit
    stf_peak_map = {}
    if stf_results:
        for r in stf_results:
            name = r.get("unit_name")
            if name:
                stf_peak_map[name] = max(stf_peak_map.get(name, 0), r.get("expected_seats", 0))

    rows = []
    for a in sorted(allocs, key=lambda x: x.seat_gap):
        u = unit_map.get(a.unit_name)
        ra = rto_map.get(a.unit_name)
        bd = breach_map.get(a.unit_name)
        cid = cluster_map.get(a.unit_name)

        gap_pct = a.seat_gap / a.effective_demand_seats if a.effective_demand_seats > 0 else 0
        risk = _risk_level(gap_pct, a.fragmentation_score)
        breach_prob = bd.get("breach_probability", 0) if bd else None
        breach_tier = _breach_tier(breach_prob) if breach_prob is not None else "—"
        stf_peak = stf_peak_map.get(a.unit_name, "N/A") if has_daily else "N/A"

        # Derive recommended action
        if risk == "RED" and a.seat_gap < 0:
            action = f"Add {abs(a.seat_gap)} seats immediately"
        elif risk == "AMBER" and a.seat_gap < 0:
            action = "Plan for seat increase within 3 months"
        elif breach_tier == "HIGH":
            action = "Adjust allocation — historical overflow >20% of days"
        elif a.fragmentation_score > 0.6:
            action = "Consolidate floors to reduce fragmentation"
        elif a.seat_gap > 30 and risk == "GREEN":
            action = "Right-size: consider reducing allocation"
        else:
            action = "No immediate action needed"

        rows.append({
            "Risk": risk,
            "Unit": a.unit_name,
            "Priority": (u.business_priority or "—") if u else "—",
            "Current HC": u.current_total_hc if u else "—",
            "Demand Seats": a.effective_demand_seats,
            "Allocated Seats": a.allocated_seats,
            "Gap": f"{a.seat_gap:+,}",
            "Fragmentation": f"{a.fragmentation_score:.2f}",
            "RTO Status": ra["status"] if ra else "N/A",
            "Breach Risk (Historical)": breach_tier,
            "STF Forecast Peak": stf_peak,
            "Cluster Group": f"Group {cid}" if cid is not None else "—",
            "Recommended Action": action,
        })

    df = pd.DataFrame(rows)
    _write_df(writer, "Unit Risk Register", df, color_col="Risk",
              color_fn=lambda v: _risk_fill(v))


def _write_scenario_comparison(writer, matrix_results):
    if not matrix_results:
        _write_df(writer, "Scenario Comparison", None,
                  note_if_empty="Run Scenario Comparison Matrix in What-If Analysis to populate this sheet.")
        return

    obj_labels = {
        "optimal_placement": "Optimal Placement",
        "rto_based": "RTO-Based",
        "rto_whatif": "What-If RTO",
    }
    rows = []
    for r in matrix_results:
        rows.append({
            "Rank": r.get("rank", "—"),
            "Mode": obj_labels.get(r.get("objective", ""), r.get("objective", "—")),
            "RTO (d/wk)": r.get("rto_mandate", "—"),
            "Alloc %": f"{r['alloc_pct']:.0%}" if r.get("alloc_pct") is not None else "N/A",
            "Demand": r.get("demand", "—"),
            "Capacity": r.get("capacity", "—"),
            "Headroom": r.get("headroom", "—"),
            "Units at Risk": r.get("units_at_risk", "—"),
            "Avg Fragmentation": r.get("avg_fragmentation", "—"),
            "Seats Saved": r.get("seats_saved", "—"),
            "Composite Score": f"{r.get('composite_score', 0):.3f}",
        })

    df = pd.DataFrame(rows)
    _write_df(writer, "Scenario Comparison", df, color_col="Rank",
              color_fn=lambda v: _fill(GREEN_F) if str(v) == "1" else _fill(GREY_F))


# ── Main entry point ──────────────────────────────────────────────────────────

def generate_holistic_report(
    scenario,
    floors: list,
    units: list,
    att_map: dict,
    rule_config: dict,
    daily_df=None,
    unit_names: Optional[List[str]] = None,
    stf_results: Optional[list] = None,
    alert_days: Optional[list] = None,
    dow_df=None,
    conflict: Optional[dict] = None,
    peak_data: Optional[list] = None,
    breach_data: Optional[list] = None,
    clusters: Optional[list] = None,
    matrix_results: Optional[list] = None,
) -> bytes:
    """
    Generate a holistic 9-sheet Excel workbook covering all domains.
    Returns bytes for st.download_button.
    """
    has_daily = daily_df is not None and not daily_df.empty
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Create placeholder so openpyxl workbook is initialised
        pd.DataFrame().to_excel(writer, sheet_name="_init_", index=False)

        _write_cover(writer, scenario, floors, units, rule_config, has_daily)
        _write_executive_summary(writer, scenario, floors, units, rule_config,
                                  stf_results, alert_days, breach_data, clusters,
                                  conflict, matrix_results, has_daily)
        _write_scenario_allocation(writer, scenario, units, att_map, rule_config)
        _write_floor_intelligence(writer, scenario, floors)
        _write_stf_and_breach(writer, scenario, stf_results, alert_days, breach_data, has_daily)
        _write_peak_day_ops(writer, scenario, peak_data, conflict, floors, has_daily, clusters)
        _write_demand_patterns(writer, dow_df, clusters, has_daily)
        _write_unit_risk_register(writer, scenario, units, att_map, rule_config,
                                   breach_data, stf_results, clusters, has_daily)
        _write_scenario_comparison(writer, matrix_results)

        # Remove init placeholder sheet
        if "_init_" in writer.book.sheetnames:
            del writer.book["_init_"]

        # Set Cover as active sheet
        if "Cover" in writer.book.sheetnames:
            writer.book.active = writer.book["Cover"]

    return output.getvalue()
